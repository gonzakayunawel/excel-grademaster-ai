import openpyxl
from io import BytesIO

def extraer_todas_las_celdas(archivo_bytes: bytes, items: list[dict]) -> dict[str, dict]:
    """
    Extrae fórmulas y valores calculados para cada ítem de la rúbrica.

    Parámetros:
        archivo_bytes: contenido del .xlsx en bytes.
        items: lista de dicts con 'hoja_objetivo' y 'celda_objetivo'.

    Retorna:
        {"Hoja1!B12": {"formula": ..., "valor": ..., "error": str|None}}
    """
    wb_f = openpyxl.load_workbook(BytesIO(archivo_bytes), data_only=False)
    wb_v = openpyxl.load_workbook(BytesIO(archivo_bytes), data_only=True)

    resultados = {}
    for item in items:
        hoja = item["hoja_objetivo"]
        celda = item["celda_objetivo"]
        es_fc = item.get("es_formato_condicional", 0) == 1
        rango_fc = item.get("rango_esperado")
        
        # Si es formato condicional, la "celda" de referencia puede ser el rango, 
        # o usamos la clave estándar si "celda_objetivo" sigue siendo un identificador
        clave = f"{hoja}!{celda}"

        if hoja not in wb_f.sheetnames:
            hojas_disponibles = ", ".join(wb_f.sheetnames)
            resultados[clave] = {
                "formula": None,
                "valor": None,
                "error": f"La hoja '{hoja}' no existe en el archivo. Hojas disponibles: {hojas_disponibles}"
            }
            continue

        ws_f = wb_f[hoja]
        ws_v = wb_v[hoja]
        
        if es_fc:
            try:
                # Buscar reglas de formato condicional
                reglas_encontradas = []
                # conditional_formatting contiene las reglas. Iteramos por sus rangos (sqref)
                for sqref, rules in ws_f.conditional_formatting._cf_rules.items():
                    for rule in rules:
                        # Extraer formulas
                        formulas = rule.formula if rule.formula else []
                        
                        # Extraer color (fill)
                        color_hex = None
                        if rule.dxfId is not None and wb_f.styles.differential_styles.elements:
                            dxf = wb_f.styles.differential_styles.elements[rule.dxfId]
                            if dxf.fill and dxf.fill.bgColor and dxf.fill.bgColor.rgb:
                                color_hex = dxf.fill.bgColor.rgb
                        
                        reglas_encontradas.append({
                            "rango": str(sqref),
                            "tipo": rule.type,
                            "formulas": formulas,
                            "color_hex": color_hex
                        })
                
                # Para la UI, devolvemos las reglas encontradas serializadas (o procesadas en scoring)
                resultados[clave] = {
                    "formula": None, # No aplica directamente a una celda
                    "valor": None,
                    "reglas_fc": reglas_encontradas,
                    "error": None
                }
            except Exception as e:
                resultados[clave] = {
                    "formula": None,
                    "valor": None,
                    "error": f"Error al extraer formato condicional en hoja {hoja}: {str(e)}"
                }
        else:
            try:
                resultados[clave] = {
                    "formula": ws_f[celda].value,
                    "valor": ws_v[celda].value,
                    "error": None
                }
            except Exception as e:
                resultados[clave] = {
                    "formula": None,
                    "valor": None,
                    "error": f"Error al leer celda {celda}: {str(e)}"
                }

    return resultados
